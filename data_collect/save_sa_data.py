import xlwt,xlrd
from data_collect.login_emos import *
import pandas as pd
import json
from datetime import datetime,timedelta
import urllib.parse
from openpyxl import load_workbook

def get_sa_city_data(s,date):
    url = 'http://10.53.160.65/jksdev_190_171_8050/sa/api/develop/loadDevelopDayData'
    fileurl = '../data/sa_city_daily.xls'
    if not os.path.exists(fileurl):
        workbook = xlwt.Workbook(encoding='ascii')
        worksheet = workbook.add_sheet("SAE下载速率")
        worksheet.write(0, 0, "日期")
        worksheet.write(0, 1, "地市")
        worksheet.write(0, 2, "5G终端用户数")
        worksheet.write(0, 3, "5G终端用户占比")
        worksheet.write(0, 4, "5G软开关开启")
        worksheet.write(0, 5, "5G软开关开启占比")
        worksheet.write(0, 6, "SA注册请求用户数")
        worksheet.write(0, 7, "SA注册请求增长")
        worksheet.write(0, 8, "SA注册成功用户数")
        worksheet.write(0, 9, "SA活跃用户数")
        worksheet.write(0, 10, "SA活跃增长数")
        worksheet.write(0, 11, "NSA活跃用户")
        worksheet.write(0, 12, "SA流量(GB)")
        worksheet.write(0, 13, "SA流量增长(GB)")
        worksheet.write(0, 14, "NSA流量(GB)")
        workbook.save(fileurl)
    df = pd.read_excel(fileurl)
    columns = ["日期","地市","5G终端用户数","5G终端用户占比","5G软开关开启","5G软开关开启占比","SA注册请求用户数","SA注册请求增长"
               ,"SA注册成功用户数","SA活跃用户数","SA活跃增长数","NSA活跃用户","SA流量(GB)","SA流量增长(GB)","NSA流量(GB)"]
    if date not in list(df['日期'].astype('str')):
        payload = {"endTime": date, "key": "sa_user_active_cnt", "queryTime": date, "startTime": "20210126",
                   "type": "developCity"}
        r = s.post(url, json=payload)
        js = json.loads(r.text)
        citydatas = js['data']['rows']
        if len(citydatas):
            for citydata in citydatas:
                if len(citydata) >=13:
                    citydataform =[[date,citydata['city_name'],citydata['user_5gterm_count'],citydata['pct_5gterm'],citydata['term_5g_open_cnt'],
                                    citydata['pct_5gopen'],citydata['sa_user_cnt'],0,citydata['sa_user_re_succ_cnt'],citydata['sa_user_active_cnt'],
                                    0,citydata['nsa_user_cnt'],round(citydata['sa_flow_gb'],2),0,round(citydata['nsa_flow_gb'],2)]]
                    #print(columns)
                    #print(citydataform)
                    df1 = pd.DataFrame(citydataform,columns=columns)
                    df = df.append(df1)
    df.to_excel(fileurl,index=None)

def get_sa_xm_data(s,date):
    url = "http://10.53.160.65/jksdev_190_171_8050/sa/api/develop/loadDevelopDayData"
    fileurl = '../data/sa_xm_daily.xls'
    columns = ["日期", "区县", "5G终端用户数", "5G终端用户占比", "5G软开关开启", "5G软开关开启占比", "SA注册请求用户数", "SA注册请求增长",
               "SA注册成功用户数", "SA活跃用户数", "SA活跃增长数", "NSA活跃用户", "SA流量(GB)", "SA流量增长(GB)", "NSA流量(GB)"]
    if not os.path.exists(fileurl):
        df = pd.DataFrame(columns=columns)
        df.to_excel(fileurl,sheet_name='SA厦门分区指标',index=None)
    df = pd.read_excel(fileurl)
    if date not in list(df['日期'].astype('str')):
        payload = {"city": "厦门", "county": "all", "endTime": date, "key": "sa_user_active_cnt",
                   "queryTime": date, "startTime": "20210126", "type": "developCounty"}
        r = s.post(url, json=payload)
        js = json.loads(r.text)
        countrydatas = js['data']['rows']
        if len(countrydatas):
            for countrydata in countrydatas:
                if len(countrydata) >= 13:
                    countrydataform = [[date, countrydata['county_name'], countrydata['user_5gterm_count'],
                                     countrydata['pct_5gterm'], countrydata['term_5g_open_cnt'],
                                     countrydata['pct_5gopen'], countrydata['sa_user_cnt'], 0,
                                     countrydata['sa_user_re_succ_cnt'], countrydata['sa_user_active_cnt'],
                                     0, countrydata['nsa_user_cnt'],
                                     0, 0,round(countrydata['nsa_flow_gb'], 2)]]
                    df1 = pd.DataFrame(countrydataform, columns=columns)
                    df = df.append(df1)
    df.to_excel(fileurl, index=None)

def get_xm_5G_visitrate(s,date):
    cityname = '厦门'
    url = 'http://10.53.160.65/jksdev_29_203_8080/jkIdcZb/sa/getDataTerminalCounty?date='+date+'&city='+urllib.parse.quote(cityname)
    r = s.get(url)
    #print(r.text)
    fileurl = '../data/sa_xm_visitrate.xls'
    columns = ["日期", "区县","登网率"]
    if not os.path.exists(fileurl):
        df = pd.DataFrame(columns=columns)
        df.to_excel(fileurl,sheet_name='厦门5G分区登网率',index=None)
    df = pd.read_excel(fileurl)
    if date not in list(df['日期'].astype('str')):
        r = s.get(url)
        js = json.loads(r.text)
        if len(js):
            for countrydata in js:
                countrydataform = [[date,countrydata['county_name'],countrydata['terminal_visit_rate']]]
                df1 = pd.DataFrame(countrydataform, columns=columns)
                df = df.append(df1)
    df.to_excel(fileurl, index=None)

def get_city_5G_visitrate(s,date):
    cityname = '全省'
    url = 'http://10.53.160.65/jksdev_29_203_8080/jkIdcZb/sa/getDataTerminalCity?date='+date+'&city='+urllib.parse.quote(cityname)
    r = s.get(url)
    #print(r.text)
    fileurl = '../data/sa_city_visitrate.xls'
    columns = ["日期", "地市","登网率"]
    if not os.path.exists(fileurl):
        df = pd.DataFrame(columns=columns)
        df.to_excel(fileurl,sheet_name='地市5G登网率',index=None)
    df = pd.read_excel(fileurl)
    if date not in list(df['日期'].astype('str')):
        r = s.get(url)
        js = json.loads(r.text)
        if len(js):
            for countrydata in js:
                countrydataform = [[date,countrydata['city_name'],countrydata['terminal_visit_rate']]]
                df1 = pd.DataFrame(countrydataform, columns=columns)
                df = df.append(df1)
    df.to_excel(fileurl, index=None)

def get_city_epsfb_delay(s,date):
    url = "http://10.53.160.65/jksdev_190_171_8050/sa/api/eps/loadEPSDayData"
    fileurl = '../data/epsfb_city_daily.xls'
    columns = ["日期", "地市", "回落时延", "E2V时延"]
    if not os.path.exists(fileurl):
        df = pd.DataFrame(columns=columns)
        df.to_excel(fileurl,sheet_name='地市EPSFB指标',index=None)
    df = pd.read_excel(fileurl)
    if date not in list(df['日期'].astype('str')):
        payload = {"endTime": date, "key": "reg_suc_usr_cnt", "queryTime": date, "startTime": "20210122",
                   "type": "epsCity"}
        r = s.post(url, json=payload)
        js = json.loads(r.text)
        citydatas = js['data']['rows']
        if len(citydatas):
            for citydata in citydatas:
                if "epsfb_resp_delay_avg" in citydata and "epsfb_build_net_delay_E2V" in citydata:
                    citydataform = [[date, citydata['REGION_NAME_NR'],  round(citydata['epsfb_resp_delay_avg'],0),
                                     round(citydata['epsfb_build_net_delay_E2V'],0)]]
                    df1 = pd.DataFrame(citydataform, columns=columns)
                    df = df.append(df1)
    df.to_excel(fileurl, index=None)

def epsfb_analysis_city_data(s,date):
    url = "http://10.53.160.65/jksdev_190_171_8050/sa/api/eps/loadEPSDayData"

    fileurl = '../data/epsfb_analysis_city_daily.xls'
    columns = ["日期", "地市", "E2EV接续时延", "E2EV接通次数", "E2EV超8s次数",
               "E2E接续时延", "E2E接通次数", "E2E超8s次数","E2V接续时延", "E2V接通次数", "E2V超8s次数"]
    if not os.path.exists(fileurl):
        df = pd.DataFrame(columns=columns)
        df.to_excel(fileurl, sheet_name='EPSFB 分析数据', index=None)
    df = pd.read_excel(fileurl)
    if date not in list(df['日期'].astype('str')):
        payload = {"endTime": date, "startTime": date,"type": "epsCityDelay"}
        r = s.post(url, json=payload)
        #print(r.text)
        js = json.loads(r.text)
        citydatas = js['data']['rows']
        if len(citydatas):
            for citydata in citydatas:
                citydataform = [[date, citydata['region_name_nr'], round(citydata['mo_epsfb_180_delay_avg_e2e_e2v'],0), citydata['mo_epsfb_180_e2e_e2v'],
                                 citydata['mo_epsfb_delay_gt8s_e2e_e2v'],round(citydata['mo_epsfb_180_delay_avg_e2e'],0), citydata['mo_epsfb_180_e2e'],
                                 citydata['mo_epsfb_delay_gt8s_e2e'],round(citydata['mo_epsfb_180_delay_avg_e2v'],0),citydata['mo_epsfb_180_e2v'], citydata['mo_epsfb_delay_gt8s_e2v']]]
                df1 = pd.DataFrame(citydataform, columns=columns)
                df = df.append(df1)
    df.to_excel(fileurl, index=None)

def get_sa_month_data(s):
    days = list(range(1, 31))
    days.reverse()
    # print(days)
    for n in days:
        today = datetime.now()
        delta = timedelta(days=n)
        date = (today - delta).strftime('%Y%m%d')
        #print(date)
        if date=='20210503' or date == '20210522':
            continue
        #print(date)
        get_sa_city_data(s, date)
        get_sa_xm_data(s, date)
        get_city_5G_visitrate(s,date)
        get_xm_5G_visitrate(s, date)
        get_city_epsfb_delay(s, date)
        epsfb_analysis_city_data(s, date)
    fileurl = '../data/sa_city_daily.xls'
    df = pd.read_excel(fileurl)
    df = df.drop_duplicates()
    df.to_excel(fileurl, sheet_name='SA地市指标',index=None)
    fileurl = '../data/sa_xm_daily.xls'
    df = pd.read_excel(fileurl)
    df = df.drop_duplicates()
    df.to_excel(fileurl, sheet_name='SA厦门分区指标', index=None)
    fileurl = '../data/sa_city_visitrate.xls'
    df = pd.read_excel(fileurl)
    df = df.drop_duplicates()
    df.to_excel(fileurl, sheet_name='地市5G登网率', index=None)
    fileurl = '../data/sa_xm_visitrate.xls'
    df = pd.read_excel(fileurl)
    df = df.drop_duplicates()
    df.to_excel(fileurl, sheet_name='厦门5G分区登网率', index=None)
    fileurl = '../data/epsfb_city_daily.xls'
    df = pd.read_excel(fileurl)
    df = df.drop_duplicates()
    df.to_excel(fileurl, sheet_name='地市EPSFB指标', index=None)
    fileurl = '../data/epsfb_analysis_city_daily.xls'
    df = pd.read_excel(fileurl)
    df = df.drop_duplicates()
    df.to_excel(fileurl, sheet_name='EPSFB 分析数据', index=None)

def combine_sheet():
    combine_url = "../data/combine.xlsx"
    writer = pd.ExcelWriter(combine_url)
    fileurl = '../data/sa_city_daily.xls'
    df = pd.read_excel(fileurl)
    df.to_excel(writer,sheet_name="SA地市指标",index=None)
    fileurl = '../data/sa_xm_daily.xls'
    df = pd.read_excel(fileurl)
    df.to_excel(writer, sheet_name="SA厦门分区指标",index=None)
    fileurl = '../data/sa_city_visitrate.xls'
    df = pd.read_excel(fileurl)
    df.to_excel(writer, sheet_name="地市5G登网率", index=None)
    fileurl = '../data/sa_xm_visitrate.xls'
    df = pd.read_excel(fileurl)
    df.to_excel(writer, sheet_name="厦门5G分区登网率",index=None)
    writer.save()
    writer.close()
    employee_data_process()

#以下函数实现对已存在工作簿数据修改，而不覆盖工作簿文件,工作簿中原sheet数据不会被破坏
def combine_existdata_sheet():
    combine_url = "../data/combine.xlsx"
    writer = pd.ExcelWriter(combine_url,engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    fileurl = '../data/sa_xm_daily.xls'
    df = pd.read_excel(fileurl)
    df.to_excel(excel_writer=writer,sheet_name="SA厦门分区指标",index=None)
    writer.save()
    writer.close()

def employee_data_process():
    combine_url = "../data/employee_data.xlsx"
    writer = pd.ExcelWriter(combine_url)
    fileurl = '../data/员工数据.xls'
    df = pd.read_excel(fileurl,keep_default_na=False)
    df0 = df.loc[(df['是否使用5GSA网络']=='否') & (df['是否使用5G终端']=='是') & (df['员工所属中心']!='')][['员工姓名','员工号码','员工部门','员工所属中心']]
    df0.to_excel(writer, sheet_name="未开SA的员工", index=None)
    df1 = df0.groupby("员工部门").agg({"员工号码":"count"}).reset_index()
    df1.to_excel(writer, sheet_name="未开SA部门统计", index=None)
    df2 = df.loc[(df['是否使用5G终端'] == '是') & (df['员工所属中心']!='')].groupby("员工部门").agg({"员工号码":"count"}).reset_index()
    df2.to_excel(writer, sheet_name="部门5G终端统计", index=None)
    writer.save()
    writer.close()




if __name__ == '__main__':

    s = login_pp('huangningcheng', 'Asdf.202103')
    get_sa_month_data(s)
    combine_sheet()



    #print(urllib.parse.quote("全省"))
    #print(urllib.parse.unquote('%E5%85%A8%E7%9C%81'))
    #employee_data_process()
